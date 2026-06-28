"""Generic fixture import from JSON/CSV rows (per competition)."""
from __future__ import annotations

import csv
import io
import json
import uuid
from dataclasses import dataclass, field
from typing import Any

from app.services.worldcup_loader import (
    WORLDCUP_LEAGUE_ID,
    WORLDCUP_LEAGUE_NAME,
    WORLDCUP_SEASON,
    _parse_match_datetime,
    _flag_url,
    _team_known,
)


CSV_COLUMNS = {
    "external_id": ("external_id", "num", "id"),
    "date": ("date", "fecha"),
    "time": ("time", "hora"),
    "team1": ("team1", "home_team", "local"),
    "team2": ("team2", "away_team", "visitante"),
    "round": ("round", "jornada", "fase"),
    "ground": ("ground", "venue", "estadio"),
    "group": ("group", "group_label", "grupo"),
    "home_logo_url": ("home_logo_url",),
    "away_logo_url": ("away_logo_url",),
    "status": ("status", "estado"),
    "home_score": ("home_score", "goles_local"),
    "away_score": ("away_score", "goles_visitante"),
    "betting_open": ("betting_open",),
}


@dataclass
class ImportRowError:
    row: int
    message: str


@dataclass
class ImportPreview:
    ok: bool
    records: list[dict[str, Any]] = field(default_factory=list)
    errors: list[ImportRowError] = field(default_factory=list)


def _pick(row: dict[str, str], *keys: str) -> str | None:
    lowered = {k.strip().lower(): v for k, v in row.items()}
    for key in keys:
        val = lowered.get(key.lower())
        if val is not None and str(val).strip() != "":
            return str(val).strip()
    return None


def _map_json_match(match: dict[str, Any], index: int, competition_id: uuid.UUID) -> dict[str, Any]:
    external_id = match.get("num") or match.get("external_id") or (index + 1)
    home = match["team1"]
    away = match["team2"]
    group_label = match.get("group")
    return {
        "external_id": int(external_id),
        "home_team": home,
        "away_team": away,
        "home_logo_url": match.get("home_logo_url") or _flag_url(home),
        "away_logo_url": match.get("away_logo_url") or _flag_url(away),
        "league_name": match.get("league_name") or WORLDCUP_LEAGUE_NAME,
        "league_id": int(match.get("league_id") or WORLDCUP_LEAGUE_ID),
        "league_logo_url": match.get("league_logo_url"),
        "match_date": _parse_match_datetime(match["date"], match["time"]),
        "status": match.get("status") or "scheduled",
        "home_score": match.get("home_score"),
        "away_score": match.get("away_score"),
        "round": match.get("round"),
        "group_name": group_label,
        "group_label": group_label,
        "venue": match.get("ground") or match.get("venue"),
        "season": int(match.get("season") or WORLDCUP_SEASON),
        "betting_open": match.get("betting_open", _team_known(home) and _team_known(away)),
        "competition_id": competition_id,
    }


def parse_json_fixtures(content: str | bytes, competition_id: uuid.UUID) -> ImportPreview:
    preview = ImportPreview(ok=True)
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        preview.ok = False
        preview.errors.append(ImportRowError(row=0, message=str(e)))
        return preview
    matches = data.get("matches") if isinstance(data, dict) else data
    if not isinstance(matches, list):
        preview.ok = False
        preview.errors.append(ImportRowError(row=0, message="Expected { matches: [...] }"))
        return preview
    for i, match in enumerate(matches):
        try:
            preview.records.append(_map_json_match(match, i, competition_id))
        except Exception as e:
            preview.ok = False
            preview.errors.append(ImportRowError(row=i + 1, message=str(e)))
    return preview


def parse_csv_fixtures(content: str, competition_id: uuid.UUID) -> ImportPreview:
    preview = ImportPreview(ok=True)
    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        preview.ok = False
        preview.errors.append(ImportRowError(row=0, message="CSV has no header row"))
        return preview
    for i, row in enumerate(reader):
        try:
            match = {
                "external_id": _pick(row, *CSV_COLUMNS["external_id"]),
                "date": _pick(row, *CSV_COLUMNS["date"]),
                "time": _pick(row, *CSV_COLUMNS["time"]),
                "team1": _pick(row, *CSV_COLUMNS["team1"]),
                "team2": _pick(row, *CSV_COLUMNS["team2"]),
                "round": _pick(row, *CSV_COLUMNS["round"]),
                "ground": _pick(row, *CSV_COLUMNS["ground"]),
                "group": _pick(row, *CSV_COLUMNS["group"]),
                "home_logo_url": _pick(row, *CSV_COLUMNS["home_logo_url"]),
                "away_logo_url": _pick(row, *CSV_COLUMNS["away_logo_url"]),
                "status": _pick(row, *CSV_COLUMNS["status"]),
            }
            hs = _pick(row, *CSV_COLUMNS["home_score"])
            aw = _pick(row, *CSV_COLUMNS["away_score"])
            if hs is not None:
                match["home_score"] = int(hs)
            if aw is not None:
                match["away_score"] = int(aw)
            bo = _pick(row, *CSV_COLUMNS["betting_open"])
            if bo is not None:
                match["betting_open"] = bo.lower() in ("1", "true", "yes", "si", "sí")
            for req in ("date", "time", "team1", "team2", "round", "ground"):
                if not match.get(req if req != "ground" else "ground"):
                    raise ValueError(f"Missing required field: {req}")
            ext = match.get("external_id") or str(i + 1)
            match["num"] = int(ext)
            preview.records.append(_map_json_match(match, i, competition_id))
        except Exception as e:
            preview.ok = False
            preview.errors.append(ImportRowError(row=i + 2, message=str(e)))
    return preview


def parse_xlsx_fixtures(content: bytes, competition_id: uuid.UUID) -> ImportPreview:
    """Parse first sheet of an Excel workbook as CSV-equivalent rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        preview = ImportPreview(ok=False)
        preview.errors.append(ImportRowError(row=0, message="openpyxl not installed"))
        return preview

    preview = ImportPreview(ok=True)
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        preview.ok = False
        preview.errors.append(ImportRowError(row=0, message="XLSX sheet is empty"))
        return preview
    fieldnames = [str(h).strip() if h is not None else "" for h in header]
    lines = [",".join(fieldnames)]
    for row in rows:
        cells = []
        for val in row:
            if val is None:
                cells.append("")
            else:
                text = str(val).replace('"', '""')
                cells.append(f'"{text}"' if "," in text else text)
        lines.append(",".join(cells))
    csv_text = "\n".join(lines)
    return parse_csv_fixtures(csv_text, competition_id)
